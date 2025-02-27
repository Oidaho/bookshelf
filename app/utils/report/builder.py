from typing import Any, Dict, List

from openpyxl.worksheet.worksheet import Worksheet

from .styles import CellStyle


class ReportBuilder:
    """Класс-билдер xslx отчета. Предоставляет методы для взаимодействия с листом ws."""

    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet
        self.ws.title = "Отчет по срокам возврата"
        self.current_row = 1

    def add_row(self, data: List[Any], style: CellStyle = None) -> None:
        """Добавляет строку в текущий лист, повышая индекс текущей строки.

        Args:
            data (List[Any]): Вектор колонок с данными.
            style (CellStyle, optional): Стили ячеек, которые занимает вектор. Defaults to None.
        """
        self.ws.append(data)
        if style:
            self._apply_style_to_row(self.current_row, style)
        self.current_row += 1

    def merge_cells(self, start_col: int, end_col: int) -> None:
        """Соединяет указаные ячейки в текущей строке в одну.

        Args:
            start_col (int): Начальный столбец (ячейка).
            end_col (int): Конечный столбец (ячейка).
        """
        self.ws.merge_cells(
            start_row=self.current_row - 1,
            end_row=self.current_row - 1,
            start_column=start_col,
            end_column=end_col,
        )

    def set_row_height(self, height: int) -> None:
        """Устанавливает высоту строки в пикселях.

        Args:
            height (int): Высота строки в пикселях.
        """
        self.ws.row_dimensions[self.current_row - 1].height = height

    def set_columns_width(self, config: Dict[str, int]) -> None:
        """Устанавливает ширину колонок в пикселях.

        Args:
            config (Dict[str, int]): Словарь соответствий колонка-размер.
        """
        for col, width in config.items():
            self.ws.column_dimensions[col].width = width

    def _apply_style_to_row(self, row: int, style: CellStyle) -> None:
        """Применяет стили ко всей строке.

        Args:
            row (int): Номер строки.
            style (CellStyle): Стили строки.
        """
        for col in range(1, self.ws.max_column + 1):
            style.apply(self.ws.cell(row=row, column=col))
